# -*- coding: utf-8 -*-

import getopt
import datetime
import pandas as pd

import paramiko
import psycopg2
import csv
import sys
import os

import xlrd
import xlwt
from xlwt import Workbook
import webservice as ws
import unicodedata
import string

import cx_Oracle
from datetime import datetime


from openpyxl import Workbook
#from openpyxl.compat import range
from openpyxl.utils import get_column_letter


tab_nat=[['AA', 'APATRIDE', 'APATRIDE'], ['AD', 'ANDORRE', 'ANDORRANE'], 
 ['AE', 'EMIRATS ARABES UNIS', 'EMIRIENNE'], 
 ['AF', 'AFGHANISTAN', 'AFGHANE'], 
 ['AG', 'ANTIGUA-ET-BARBUDA', 'ANTIGUAISE ET BARBUDIENNE'], 
 ['AL', 'ALBANIE', 'ALBANAISE'], 
 ['AM', 'ARMENIE', 'ARMENIENNE'], 
 ['AO', 'ANGOLA', 'ANGOLAISE'], 
 ['AR', 'ARGENTINE', 'ARGENTINE'], 
 ['AT', 'AUTRICHE', 'AUTRICHIENNE'], 
 ['AU', 'AUSTRALIE', 'AUSTRALIENNE'], 
 ['AZ', 'AZERBAIDJAN', 'AZERBAIDJANAISE'], 
 ['BA', 'BOSNIE-HERZEGOVINE', 'BOSNIAQUE'], 
 ['BB', 'BARBADE', 'BARBADIENNE'], 
 ['BD', 'BANGLADESH', 'BANGLADAISE'], 
 ['BE', 'BELGIQUE', 'BELGE'], 
 ['BF', 'BURKINA', 'BURKINABE'], 
 ['BG', 'BULGARIE', 'BULGARE'], 
 ['BH', 'BAHREIN', 'BAHREINIENNE'], 
 ['BI', 'BURUNDI', 'BURUNDAISE'], 
 ['BJ', 'BENIN', 'BENINOISE'], 
 ['BM', 'BERMUDES', 'BERMUDIENNE'], 
 ['BN', 'BRUNEI', 'BRUNEIENNE'], 
 ['BO', 'BOLIVIE', 'BOLIVIENNE'], 
 ['BR', 'BRESIL', 'BRESILIENNE'], 
 ['BS', 'BAHAMAS', 'BAHAMIENNE'], 
 ['BT', 'BHOUTAN', 'BHOUTANAISE'], 
 ['BW', 'BOTSWANA', 'BOTSWANEENNE'], 
 ['BY', 'BIELORUSSIE', 'BIELORUSSE'], 
 ['BZ', 'BELIZE', 'BELIZIENNE'], 
 ['CA', 'CANADA', 'CANADIENNE'], 
 ['CC', 'ILES COOK', 'COOK ISLANDER'], 
 ['CD', 'REPUBLIQUE DEMOCRATIQUE DU CONGO', 'CONGOLAISE'], 
 ['CF', 'REPUBLIQUE CENTRAFRICAINE', 'CENTRAFRICAINE'], 
 ['CG', 'CONGO', 'CONGOLAISE'], 
 ['CH', 'SUISSE', 'SUISSE'], 
 ['CI', "COTE D'IVOIRE", 'IVOIRIENNE'], 
 ['CK', 'ILES COOK', 'COOK ISLANDER'], 
 ['CL', 'CHILI', 'CHILIENNE'], 
 ['CM', 'CAMEROUN', 'CAMEROUNAISE'], 
 ['CN', 'CHINE', 'CHINOISE'], 
 ['CO', 'COLOMBIE', 'COLOMBIENNE'], 
 ['CR', 'COSTA RICA', 'COSTARICAINE'], 
 ['CS', 'SERBIE', 'SERBE'], 
 ['CU', 'CUBA', 'CUBAINE'], 
 ['CV', 'CAP-VERT', 'CAP-VERDIENNE'], 
 ['CY', 'CHYPRE', 'CHYPRIOTE'], 
 ['CZ', 'REPUBLIQUE TCHEQUE', 'TCHEQUE'], 
 ['DE', 'ALLEMAGNE', 'ALLEMANDE'], 
 ['DJ', 'DJIBOUTI', 'DJIBOUTIENNE'], 
 ['DK', 'DANEMARK', 'DANOISE'], 
 ['DM', 'DOMINIQUE', 'DOMINIQUAISE'], 
 ['DO', 'REPUBLIQUE DOMINICAINE', 'DOMINICAINE'], 
 ['DZ', 'ALGERIE', 'ALGERIENNE'], 
 ['EC', 'EQUATEUR', 'EQUATORIENNE'], 
 ['EE', 'ESTONIE', 'ESTONIENNE'], 
 ['EG', 'EGYPTE', 'EGYPTIENNE'], 
 ['ER', 'ERYTHREE', 'ERYTHREENNE'], 
 ['ES', 'ESPAGNE', 'ESPAGNOLE'], 
 ['ET', 'ETHIOPIE', 'ETHIOPIENNE'], 
 ['FI', 'FINLANDE', 'FINLANDAISE'], 
 ['FJ', 'FIDJI', 'FIDJIENNE'], 
 ['FM', 'MICRONESIE', 'MICRONESIENNE'], 
 ['FO', 'ILES FEROE', 'FERINGIENE'], 
 ['FR', 'FRANCE', 'FRANCAISE'], 
 ['GA', 'GABON', 'GABONAISE'], 
 ['GB', 'ROYAUME-UNI', 'BRITANNIQUE'], 
 ['GD', 'GRENADE', 'GRENADIENNE'], 
 ['GE', 'GEORGIE', 'GEORGIENNE'], 
 ['GH', 'GHANA', 'GHANEENNE'], 
 ['GM', 'GAMBIE', 'GAMBIENNE'],
 ['GN', 'GUINEE', 'GUINEENNE'], 
 ['GQ', 'GUINEE EQUATORIALE', 'EQUATO-GUINEENNE'],
 ['GR', 'GRECE', 'GRECQUE'], 
 ['GT', 'GUATEMALA', 'GUATEMALTEQUE'], 
 ['GW', 'GUINEE-BISSAU', 'BISSAO-GUINEENNE'], 
 ['GY', 'GUYANA', 'GUYANIENNE'], 
 ['HK', 'HONG-KONG', 'CHINOISE'], 
 ['HN', 'HONDURAS', 'HONDURIENNE'], 
 ['HR', 'CROATIE', 'CROATE'], 
 ['HT', 'HAITI', 'HAITIENNE'], 
 ['HU', 'HONGRIE', 'HONGROISE'], 
 ['ID', 'INDONESIE', 'INDONESIENNE'],
 ['IE', 'IRLANDE', 'IRLANDAISE'],
 ['IL', 'ISRAEL', 'ISRAELIENNE'], 
 ['IN', 'INDE', 'INDIENNE'], 
 ['IQ', 'IRAQ', 'IRAQUIENNE'],
 ['IR', 'IRAN', 'IRANIENNE'], 
 ['IS', 'ISLANDE', 'ISLANDAISE'], 
 ['IT', 'ITALIE', 'ITALIENNE'], 
 ['JM', 'JAMAIQUE', 'JAMAIQUAINE'], 
 ['JO', 'JORDANIE', 'JORDANIENNE'], 
 ['JP', 'JAPON', 'JAPONAISE'], 
 ['KE', 'KENYA', 'KENYANE'], 
 ['KG', 'KIRGHIZISTAN', 'KIRGHIZE'],
 ['KH', 'CAMBODGE', 'CAMBODGIENNE'], 
 ['KI', 'KIRIBATI', 'KIRIBATIENNE'], 
 ['KM', 'COMORES', 'COMORIENNE'], 
 ['KN', 'SAINT-CHRISTOPHE-ET-NIEVES', 'KITTIENNE ET NEVICIENNE'], 
 ['KP', 'COREE DU NORD', 'NORD-COREENNE'], 
 ['KR', 'COREE DU SUD', 'SUD-COREENNE'], 
 ['KW', 'KOWEIT', 'KOWEITIENNE'],
 ['KZ', 'KAZAKHSTAN', 'KAZAKHE'], 
 ['LA', 'LAOS', 'LAOTIENNE'], 
 ['LB', 'LIBAN', 'LIBANAISE'], 
 ['LC', 'SAINTE-LUCIE', 'SAINT-LUCIENNE'],
 ['LI', 'LIECHTENSTEIN', 'LIECHTENSTEINOISE'],
 ['LK', 'SRI LANKA', 'SRI-LANKAISE'], 
 ['LR', 'LIBERIA', 'LIBERIENNE'],
 ['LS', 'LESOTHO', 'LESOTHANE'], 
 ['LT', 'LITUANIE', 'LITUANIENNE'], 
 ['LU', 'LUXEMBOURG', 'LUXEMBOURGEOISE'], 
 ['LV', 'LETTONIE', 'LETTONE'], ['LY', 'LIBYE', 'LIBYENNE'], 
 ['MA', 'MAROC', 'MAROCAINE'], 
 ['MC', 'MONACO', 'MONEGASQUE'], 
 ['MD', 'MOLDAVIE', 'MOLDAVE'], 
 ['ME', 'MONTENEGRO', 'MONTENEGRINE'], 
 ['MG', 'MADAGASCAR', 'MALGACHE'], ['MK', 'MACEDOINE', 'MACEDONIENNE'], 
 ['ML', 'MALI', 'MALIENNE'], 
 ['MM', 'MYANMAR', 'MYANMARAISE'], 
 ['MN', 'MONGOLIE', 'MONGOLE'], 
 ['MP', 'ILES MARIANNES DU NORD', 'MARIANNAISE'],
 ['MR', 'MAURITANIE', 'MAURITANIENNE'], 
 ['MT', 'MALTE', 'MALTAISE'], 
 ['MU', 'MAURICE', 'MAURICIENNE'], 
 ['MV', 'MALDIVES', 'MALDIVIENNE'], 
 ['MW', 'MALAWI', 'MALAWIENNE'],
 ['MX', 'MEXIQUE', 'MEXICAINE'], 
 ['MY', 'MALAISIE', 'MALAISIENNE'], 
 ['MZ', 'MOZAMBIQUE', 'MOZAMBICAINE'], 
 ['NA', 'NAMIBIE', 'NAMIBIENNE'], 
 ['NE', 'NIGER', 'NIGERIANE'], 
 ['NG', 'NIGERIA', 'NIGERIENNE'], 
 ['NI', 'NICARAGUA', 'NICARAGUAYENNE'],
 ['NL', 'PAYS-BAS', 'NEERLANDAISE'],
 ['NO', 'NORVEGE', 'NORVEGIENNE'], 
 ['NP', 'NEPAL', 'NEPALAISE'], 
 ['NR', 'NAURU', 'NAURUANE'], 
 ['NU', 'NIOUE', 'NIOUEENE'], 
 ['NZ', 'NOUVELLE-ZELANDE', 'NEW-ZELANDAIS'], 
 ['OM', 'OMAN', 'OMANAISE'],
 ['PA', 'PANAMA', 'PANAMEENNE'], 
 ['PE', 'PEROU', 'PERUVIENNE'], 
 ['PG', 'PAPOUASIE-NOUVELLE-GUINEE', 'PAPOUANE-NEO-GUINEENNE'], 
 ['PH', 'PHILIPPINES', 'PHILIPPINE'], 
 ['PK', 'PAKISTAN', 'PAKISTANAISE'], 
 ['PL', 'POLOGNE', 'POLONAISE'], 
 ['PN', 'ILE PITCAIRN', 'PITCAIRNAISE'], 
 ['PR', 'PORTO RICO', 'PORTO RICAINE'], 
 ['PS', 'PALESTINE', 'PALESTINIENNE'],
 ['PT', 'PORTUGAL', 'PORTUGAISE'], 
 ['PW', 'ILES PALAOS', 'RESSORTISSANTE DE PALAU'], 
 ['PY', 'PARAGUAY', 'PARAGUAYENNE'], 
 ['QA', 'QATAR', 'QATARIENNE'],
 ['RO', 'ROUMANIE', 'ROUMAINE'], 
 ['RS', 'SERBIE', 'SERBE'], 
 ['RU', 'RUSSIE', 'RUSSE'], 
 ['RW', 'RWANDA', 'RWANDAISE'], 
 ['SA', 'ARABIE SAOUDITE', 'SAOUDIENNE'],
 ['SB', 'ILES SALOMON', 'SALOMONAISE'],
 ['SC', 'SEYCHELLES', 'SEYCHELLOISE'], 
 ['SD', 'SOUDAN', 'SOUDANAISE'],
 ['SE', 'SUEDE', 'SUEDOISE'], 
 ['SG', 'SINGAPOUR', 'SINGAPOURIENNE'],
 ['SI', 'SLOVENIE', 'SLOVENE'], 
 ['SK', 'SLOVAQUIE', 'SLOVAQUE'], 
 ['SL', 'SIERRA LEONE', 'SIERRA-LEONAISE'],
 ['SM', 'SAINT-MARIN', 'SAINT-MARINAISE'], 
 ['SN', 'SENEGAL', 'SENEGALAISE'], 
 ['SO', 'SOMALIE', 'SOMALIENNE'], 
 ['SR', 'SURINAME', 'SURINAMAISE'], 
 ['ST', 'SAO TOME-ET-PRINCIPE', 'SANTOMEENNE'], 
 ['SU', 'URSS', 'RUSSE'], 
 ['SV', 'EL SALVADOR', 'SALVADORIENNE'],
 ['SY', 'SYRIE', 'SYRIENNE'], 
 ['SZ', 'SWAZILAND', 'SWAZIE'], 
 ['TD', 'TCHAD', 'TCHADIENNE'], 
 ['TG', 'TOGO', 'TOGOLAISE'],
 ['TH', 'THAILANDE', 'THAILANDAISE'], 
 ['TJ', 'TADJIKISTAN', 'TADJIKE'], 
 ['TK', 'TOKELAU', 'TOKELAUANE'], 
 ['TM', 'TURKMENISTAN', 'TURKMENE'], 
 ['TN', 'TUNISIE', 'TUNISIENNE'],
 ['TO', 'TONGA', 'TONGUIENNE'],
 ['TP', 'TIMOR ORIENTAL', 'TIMORAISE'],
 ['TR', 'TURQUIE', 'TURQUE'], 
 ['TT', 'TRINITE-ET-TOBAGO', 'TRINIDADIENNE'], 
 ['TV', 'TUVALU', 'TUVALUANE'], 
 ['TW', 'TAIWAN', 'TAIWANAIS'],
 ['TZ', 'TANZANIE', 'TANZANIENNE'],
 ['UA', 'UKRAINE', 'UKRAINIENNE'], 
 ['UG', 'OUGANDA', 'OUGANDAISE'], 
 ['US', "ETATS-UNIS D'AMERIQUE", 'AMERICAINE'], 
 ['UY', 'URUGUAY', 'URUGUAYENNE'],
 ['UZ', 'OUZBEKISTAN', 'OUZBEKE'], 
 ['VA', 'VATICAN', 'RESSORTISSANTE DU SAINT-SIEGE'], 
 ['VC', 'SAINT-VINCENT-ET-LES GRENADINES', 'SAINT-VINCENTAISE ET GRENADINE'],
 ['VE', 'VENEZUELA', 'VENEZUELIENNE'],
 ['VN', 'VIET NAM', 'VIETNAMIENNE'], 
 ['VU', 'VANUATU', 'VANUATUANE'], 
 ['WS', 'SAMOA OCCIDENTALES', 'SAMOANE'], 
 ['XH', 'TCHECOSLOVAQUIE', 'TCHECOSLOVAQUE'], 
 ['XK', 'KOSOVO', 'KOSOVAR'],
 ['YE', 'YEMEN', 'YEMENITE'], 
 ['YU', 'YOUGOSLAVIE', 'YOUGOSLAVE'], 
 ['ZA', 'AFRIQUE DU SUD', 'SUD-AFRICAINE'], 
 ['ZC', 'KOSOVO', 'KOSOVAR'],
 ['ZM', 'ZAMBIE', 'ZAMBIENNE'], 
 ['ZR', 'REPUBLIQUE DEMOCRATIQUE DU CONGO', 'CONGOLAISE'], 
 ['ZW', 'ZIMBABWE', 'ZIMBABWEENNE']]

columns_base_specifique = ['Siren', 'DenominationSociale', 'CodeGreffe', 'CodeFormeJuridique', 'LibelleFormeJuridique','NombreDeBeneficiaireEffectif',
                'AdresseSiegeSocial', 'CodePostalSiegeSocial', 'CommuneSiegeSocial', 'PaysSiegeSocial', 'CodeCommuneInsee',  
                'DetentionCapital', 'PourcentageDetentionCapital', 'DetentionDroitsDeVote', 'PourcentageDroitsDeVote',
                'AutreMoyenControle', 'RepresentantLegalEstBeneficiaireEffectif', 'DateEffetBeneficiaireEffectif', 
                'CiviliteBeneficiaireEffectif', 'NomPatronymiqueBeneficiaireEffectif','NomUsageBeneficiaireEffectif', 'PseudonymeBeneficiaireEffectif', 
                'PrenomsBeneficiaireEffectif', 'DateNaissanceBeneficiaireEffectif', 'CommuneNaissanceBeneficiaireEffectif', 
                'DepartementPaysNaissanceBeneficiaireEffectif',
                'NationaliteBeneficiaireEffectif', 'PaysNaissanceBeneficiaireEffectif', 'CodeCommuneInseeNaissanceBeneficiaireEffectif',
                'AdresseDomicileBeneficiaireEffectif', 'CodePostalDomicileBeneficiaireEffectif', 'CommuneDomicileBeneficiaireEffectif',
                'PaysDomicileBeneficiaireEffectif', 'CodeCommuneInseeDomicileBeneficiaireEffectif', 
                'TypeDeclaration','DateActe','DateDepot','CodeRetour','PremierPrenomsBeneficiaireEffectif','AutresPrenomsBeneficiaireEffectif','CodePaysNaissanceBeneficiaireEffectif']

nsfic= {'fic': 'ficheid.webserv.experian.com'}

columns_base_ax=[
'Numéro d\'article','Quantité','Unité','Prix unitaire','Montant TTC','Greffe','Service','Code GIE','SIREN demandé','Dénomination SIREN demandé','Date de commande']


def recup_fj(siren,sirens,fj):

    for i in range(len(sirens)):
        if siren==sirens[i]:           
            codefj=fj[i]
            return(codefj)
            
def civilite(civilite_str):
        
        if civilite_str == "Mme" or civilite_str == "MLLE" or civilite_str == "Mlle" :
            civilite_str="MME"
            
        return(civilite_str)
            
def replace(ligne_base, index):

    if ligne_base[index] == 0:
        ligne_base[index] = 'direct'
    elif ligne_base[index] == 1:
        ligne_base[index] = 'indirect'
    elif ligne_base[index] == 2:
        ligne_base[index] = 'non précisé'
    elif ligne_base[index] == 3:
        ligne_base[index] = 'direct et indirect'
    
def normalize(text):

   text = unicodedata.normalize('NFD', text)
   ascii = False
   chars = []
   for c in text:
       if unicodedata.combining(c) and ascii:
           continue
       chars.append(c)
       if not unicodedata.combining(c):
           ascii = c in string.ascii_letters
   text = unicodedata.normalize('NFC', ''.join(chars))
   return text.upper()
   
def recup_nat(nationalite):
    for i in range(len(tab_nat)):
        if nationalite==tab_nat[i][2]:
            return(nationalite)
    return("")
            
def recup_pays(pays):
    for i in range(len(tab_nat)):
        if tab_nat[i][1] in pays:
            pays=tab_nat[i][1]
            return(pays)
    return("")
    
def code_pays(pays):
    for i in range(len(tab_nat)):
        if tab_nat[i][1] in pays:
            codepays=tab_nat[i][0]
            return(codepays)
    return("")
    
def tr_pays(pays):
    for i in range(len(tab_nat)):
        #print(tab_nat[i][0])
        if tab_nat[i][0] in pays:
            #print(tab_nat[i][1])
            pays=tab_nat[i][1]
            return(pays)
    return("")

def base_lines(liste_sirens, fj, statut_code):

    
    liste_sirens_ok = set()
    str_in="siren IN "+str(liste_sirens).replace('[','(').replace(']',')')
    
    if statut_code=='nocode':
        str_coderetour="lb.coderetour::int=0"
    else:
        str_coderetour="((lb.coderetour::int >= 150 and lb.coderetour::int < 161 ) or lb.coderetour::int=0)"
    

    req_base_ligne = "SELECT  \
                     lb.siren,  \
                     lb.denomination,  \
                     lb.codegreffe,  \
                     lb.codeformejur,  \
                     lb.libformejuridique,  \
                     lb.nbpps_be,  \
                     lb.adrvoie_be,  \
                     lb.adrcp_be,  \
                     lb.adrcomm_be,  \
                     lb.adrpays_be,  \
                     lb.adrcodecomm_be,  \
                     lb.detcapital,  \
                     lb.pcdetcapital,  \
                     lb.detvotes,  \
                     lb.pcdetvotes,  \
                     lb.autremoyen,  \
                     lb.replegal,  \
                     lb.dteffet,  \
                     lb.ppcivilite, \
                     lb.ppnompatrim, \
                     lb.ppnommarital,  \
                     lb.pppseudonyme,  \
                     lb.ppprenoms,  \
                     lb.ppdtnaiss,  \
                     lb.ppvillenaiss,  \
                     lb.ppdeptnaiss,  \
                     lb.ppnationalite,  \
                     lb.pppaysnaiss,  \
                     lb.ppcodecomnaiss,  \
                     lb.adr_be,  \
                     lb.cp_be,  \
                     lb.comm_be,  \
                     lb.pays_be,  \
                     lb.codecomm_be,  \
                     lb.typedoc,  \
                     lb.dtacte,  \
                     lb.dtdepot,  \
                     lb.coderetour \
                     FROM lignes_beffectifs lb \
                     WHERE "+str_coderetour+" and " + str_in +" and operateur='AMITEL' order by 1 "
                     
    #print(req_base_ligne)
    cursor_pg.execute(req_base_ligne)
#
    lignes_base = []

    for row in cursor_pg.fetchall():
        ligne_base = []
        for i in range(len(row)):
            ligne_base.append(row[i])
            
        siren=ligne_base[0]
        
        codefj=recup_fj(siren,liste_sirens,fj)
        if codefj[len(codefj)-2:] == 'Uh':
            ligne_base[3]=codefj
    
        ligne_base[4]=ligne_base[4].replace('"','')
        liste_sirens_ok.add(ligne_base[0])

        if ligne_base[37]=='149':
            ligne_base[37]='0'
        
        
        if ligne_base[15]=="<!-- texte -->" or ligne_base[15]=="<!-- illisible -->" or ligne_base[15]=="<!-- schema -->":
            ligne_base[15] = 'oui'

        if ligne_base[16] == 1:
            ligne_base[16] = 'oui'
            
        if ligne_base[11] == 2 and ligne_base[12] is None:
            ligne_base[11] = ""
        if ligne_base[13] == 2 and ligne_base[14] is None:
            #print(ligne_base[15])
            ligne_base[13] = ""
        replace(ligne_base, 11)
        replace(ligne_base, 13)
        if ligne_base[17] is not None:
            ligne_base[17]=ligne_base[17].strftime("%d-%m-%Y")
            
        # suppression de NEE dans le prenom
        index=ligne_base[22].find(' NEE ')
        if index != -1:
            ligne_base[22]=ligne_base[22][:index]
            
        ligne_base[22]=ligne_base[22].replace('/',' ')
        ligne_base[18]=civilite(ligne_base[18])
        
        if ligne_base[23] is not None:
            ligne_base[23]=ligne_base[23].strftime("%d-%m-%Y")
        if ligne_base[24] is not None:
            ligne_base[24]=ligne_base[24].replace('.','')
        
        if ligne_base[35] is not None:
            ligne_base[35]=ligne_base[35].strftime("%d-%m-%Y")
        if ligne_base[36] is not None:
            ligne_base[36]=ligne_base[36].strftime("%d-%m-%Y")
        
#            print(ligne_base[26])
#            print(ligne_base[37])
        if ligne_base[26] is not None:
            nationalite=recup_nat(normalize(ligne_base[26].replace("'","").strip()))
        else:
            nationalite=""
        ligne_base[26]=nationalite
        
        if nationalite=="" and statut_code !="nocode":
            ligne_base[37]='152'
            
        if ligne_base[27] is not None:
            pays=recup_pays(normalize(str(ligne_base[27]).replace("'","").strip()))
            codepays=code_pays(normalize(str(ligne_base[27]).replace("'","").strip()))
        else:
            pays=""
            codepays=""
            
        ligne_base[27]=pays
        
        
        if pays=='FRANCE' and ligne_base[28] != "":
            ligne_base[25]=str(ligne_base[28])[0:2]
        elif pays != "":
            ligne_base[25]='99'
        else:
            ligne_base[25]=''
        
        ligne_base[25]=ligne_base[25].replace('No','')
        
        ligne_base[22]=normalize(ligne_base[22])
        
        if ligne_base[32] is not None:
            pays=recup_pays(normalize(str(ligne_base[32]).replace("'","").strip()))
        else:
            pays=""
        ligne_base[32]=pays

        
        if ligne_base[6] is None or ligne_base[6] =="":
            return_node=ws.get_ficheEntrepriseEnrichie(ligne_base[0])
            fiche_node=return_node.find('fic:ficheIdentiteEnrichie',nsfic)
            siren_ws=fiche_node.find('fic:siren',nsfic)
            if siren_ws is not None:
                adresse_node=fiche_node.find('fic:adresse',nsfic)
                if adresse_node is not None:
                    adresse_concat=adresse_node.find('fic:adresseConcat',nsfic).text
                    if adresse_concat is not None:
                        ligne_base[6]=adresse_concat.replace("'"," ")
                
                adrcp=adresse_node.find('fic:codePostal',nsfic).text
                if adrcp is not None:
                    ligne_base[7]=str(adrcp)
                    
                bureau_distributeur=adresse_node.find('fic:bureauDistributeur',nsfic).text
                if bureau_distributeur is not None:
                    ligne_base[8]=bureau_distributeur.replace("'"," ")
                    
        autresprenoms=""
        prenoms=ligne_base[22].split(' ')
        ligne_base.append(prenoms[0])
        if len(prenoms) > 1:
            for i in range(1,len(prenoms)):
                autresprenoms=autresprenoms+" "+prenoms[i]
                
        ligne_base.append(autresprenoms)
        
        ligne_base.append(codepays)
        
        lignes_base.append(ligne_base)
    
    #print(liste_sirens_ok)

    return lignes_base
    
    
def requete_eligibilite(sirens):

    init_cmd="SELECT "+str(sirens[0])+ "AS siren FROM dual \n"
    ligne_union=""
    for i in range(1,len(sirens)):
        ligne_union=ligne_union+"UNION SELECT "+str(sirens[i])+ "        FROM dual \n"
        
    CMD_SELECT=init_cmd+ligne_union
        
    #print(siren)

    SQL="select \
siren as SIREN_RECHERCHE \
, case \
when e.ent_tre_c is not null then e.ent_tre_c \
when insent_siren is not null then 'INSEE' \
else 'SIREN NON TROUVE' \
end as INSCRIPTION \
, case \
when rdo_gr_nume in (select gr_nume from e_greffe where fs_verif_greffe_ouvert(gr_nume, '01 02 03 05 51', null) = 1) then 'GTC' \
when rdo_gr_nume in (select gr_nume from e_greffe where fs_verif_greffe_ouvert(gr_nume,'20', null) = 1) then'TI/TMC' \
else null \
end as source \
,nvl(rdo_deno_nom ,insent_nomen  ) as DENOMINATION_TROUVEE \
,case \
when rdo_st_c= 'A' then 'ZZZh' \
when fo_ju_c is not null then fo_ju_c \
else inscj_c \
end  as codeFJ \
,case \
when rdo_st_c= 'A' then 'Commercant' \
when fo_ju_c is not null then fo_ju_l \
else inscj_l \
end  as FJ \
, (case \
when not(rdo_radid is null and nvl(rdo_et_ra_c,'0')='0') then 'OUI' \
when rdo_gr_nume is null and insent_cess_d is not null then 'OUI' \
else null \
end) \
as \"RADIE / CESSE ?\" \
, (case  \
when rdo_fo_ju_c in (SELECT  gpt_val            \
                FROM  G_PARM_TRAIT \
                WHERE GR_NUME        = '9999' \
                AND   GPT_RD_C       = 'IP' \
                AND   GPT_CP_C       = 'W_IP_KYC_ACTE_CONF_FJ') \
                then 'NON' \
when rdo_sir_nume in (SELECT  gpt_val \
           FROM  G_PARM_TRAIT \
           WHERE GR_NUME        = '9999' \
           AND   GPT_RD_C       = 'IP' \
           AND   GPT_CP_C       = 'W_IP_KYC_ACTE_CONF_COTEE' \
           AND   GPT_COMP_C     = 'ALL_GREFFES' \
)             then 'NON' \
when rdo_st_c in ('A','P','I') then 'NON' \
when rdo_st_c in ('B','C','D','M')  then 'OUI' \
else null \
end) \
as ELIGIBLE_RBE \
, (select distinct 'OUI' \
          from e_rcs_dossier, e_acte_actes \
         where rdo_sir_nume = e.ent_sir_nume \
           and rdo_gr_nume = aac_gr_nume \
           and rdo_mil = aac_mil \
           and rdo_st_c = aac_st_c \
           and rdo_chrono = aac_chrono \
           and AAC_ACTE_TY_C like 'BE%') as DEPOT_DIBE \
from ( "+CMD_SELECT+") liste_siren \
left join e_entreprise e on e.ent_tre_c = 'RCS'  and e.ent_sir_nume =  liste_siren.siren \
left join e_rcs_dossier r on r.rdo_gr_nume = e.ent_princ_gr_nume \
and r.rdo_mil = e.ent_princ_mil and r.rdo_st_c = e.ent_princ_st_c and r.rdo_chrono = e.ent_princ_chrono \
left join e_forme_juridique j on j.fo_ju_gr_nume = rdo_gr_nume and j.fo_ju_c = rdo_fo_ju_c \
left join e_forme_juridique_harmo jh on jh.fjh_c = rdo_fo_ju_c \
left join e_fai_der_evt_code on fde_c not in ('A5') \
and fde_c = GRFADM.fs_get_etat_procedure (ENT_PRINC_GR_NUME   , ENT_PRINC_MIL     , ENT_PRINC_ST_C,   ENT_PRINC_CHRONO    ) \
left join ins_entreprise on insent_siren = liste_siren.siren \
left join ins_cj on inscj_c = insent_cj_c \
order by SIREN_RECHERCHE" 



    return(SQL)

def eligibilite(liste_sql):
    liste_result=[]
    conn = cx_Oracle.connect('grf_infbatch/inf83batch!@GRFPRD')
    cursor = conn.cursor()
    
    cursor.execute(liste_sql)
    result = cursor.fetchall()
    for j in range(len(result)):
        liste_result.append(result[j])
    
    cursor.close()
    conn.close()

    return(liste_result)
    

columns_base = ['Siren', 'Inscription', 'Source', 'Denomination','CodeFormeJuridique','LibelleFormeJuridique','RADIE/CESSE','EligibleRBE','DepotDIBE']

optlist, args = getopt.getopt(sys.argv[1:], 'ef:s:i:d:')

liste_sirens=[]
liste_unique=set()
file='true'
entete='false'
indice=0
root_dir="."
ident='0'


for opt, arg in optlist:

    if opt in ("-f"):
        filename = arg
        file='true'
    elif opt in ("-s"):
        liste_unique.add(arg)
        file='false'
    elif opt in ("-d"):
        root_dir=arg
    elif opt in ("-i"):
        ident=arg
    elif opt in ("-e"):
        entete='true'
            
if file == 'true':
    root,ext=os.path.splitext(filename)
    if ext==".csv":
        file_siren=open(filename,'r')
        #lecture de l'entete
        if entete=='true':
            file_siren.readline()
        for ligne in file_siren:
            siren=ligne.replace('\n','').split(";")[0].zfill(9)
            
            liste_unique.add(siren)
            
    elif ext==".xls":
        wb = xlrd.open_workbook(filename)
        sh=wb.sheet_by_index(0)
        if entete=='true':
            indice=1
        for rownum in range(indice,sh.nrows):
            celltype=sh.cell_type(rownum,0)
            if  celltype==2 or celltype==1:
                try:
                    siren=str(int(sh.cell_value(rownum,0))).zfill(9)
                    if len(siren)==9:
                        liste_unique.add(siren)
                except:
                    print(sh.cell_value(rownum,0))
            

    elif ext==".xlsx":
        wb = xlrd.open_workbook(filename)
        sh=wb.sheet_by_index(0)
        if entete=='true':
            indice=1
        for rownum in range(indice,sh.nrows):
            celltype=sh.cell_type(rownum,0)
            if sh.cell_type(rownum,0) == 2:
                siren=str(int(sh.row_values(rownum)[0])).zfill(9)
                liste_unique.add(siren)
            elif sh.cell_type(rownum,0) ==1:
                siren=str(sh.row_values(rownum)[0]).replace(' ','').zfill(9)
                if siren==siren.lower() and siren==siren.upper():
                    liste_unique.add(siren)
                else:
                    print(siren +"pas un siren")
            else:
                print("problème sur type de cellule")
                    

    else:
        print("Format de fichier non reconnue")
        exit()
        
else:
    print("test unitaire")
    
liste_sirens=list(liste_unique)

liste_sirens.sort()

nb_unit=2000
lignes=[]
cmd_sql=[]

nb=len(liste_sirens)

reste=nb % nb_unit
nb=int(nb/nb_unit)
excel_names = []

if reste != 0:
    nb=nb+1
    
    
PG_CONFIG = {
'host': '79.137.30.193',
'dbname': 'DATAIFG_TEST2',
'user': 'infogreffe',
'password': '3Mg0Fs2Eg2'
}

try:
    conn_pg = psycopg2.connect(**PG_CONFIG)
    cursor_pg = conn_pg.cursor()
    
except:
    print ("I am unable to connect to the database Associes")
    exit()


cpt=0
for i in range(nb):
    liste_eligibilite_all=[]
    
    liste10000=liste_sirens[nb_unit*i:(i+1)*nb_unit]
    cmd_eligible=requete_eligibilite(liste10000)
    #cmd_sql.append(cmd_eligible)

    lignes=eligibilite(cmd_eligible)

    #create_table(liste10000)
    liste_eligibilite=[]
    for i in range(len(lignes)):
        liste_eligibilite.append(list(lignes[i]))

    liste_eligibilite_all.extend(liste_eligibilite)
    #delete_table()
    

            # création 
    book = Workbook()
     
    # création de la feuille 1
    feuil1 = book.active
 
# ajout des en-têtes

    if cpt == 0:
        for m in range(len(columns_base)):
            feuil1.cell(column=m+1,row=1,value=columns_base[m])
        
    for l in range(len(liste_eligibilite_all)):
        siren=liste_eligibilite_all[l][0]
        for k in range(len(liste_eligibilite_all[l])):
            valeur=str(liste_eligibilite_all[l][k]).replace("None","")
            if k==8 and str(liste_eligibilite_all[l][k]) != 'OUI' and str(liste_eligibilite_all[l][k-1]) == 'OUI':
                valeur='NON'
            feuil1.cell(column=k+1,row=l+2,value=valeur)

# création matérielle du fichier résultant

    if file=='true':
        root,tmpfi = os.path.split(filename)
        fname,ext=os.path.splitext(tmpfi)
        output_name=root_dir+"/"+fname+"_"+str(cpt)+"_CR.xlsx"
        cpt=cpt+1
        book.save(output_name)
        excel_names.append(output_name)
    else:
        fname=str(siren)
        output_name=root_dir+"/"+fname+"_CR.xlsx"
        book.save(output_name)

if file=='true':
    excels = [pd.ExcelFile(name) for name in excel_names]
    
    # turn them into dataframes
    frames = [x.parse(x.sheet_names[0], header=None,index_col=None) for x in excels]
    
    # delete the first row for all frames except the first
    # i.e. remove the header row -- assumes it's the first
    frames[1:] = [df[1:] for df in frames[1:]]
    
    # concatenate them..
    combined = pd.concat(frames)
    
    # write it out
    fileoutname=root_dir+"/"+fname+"_CR.xlsx"
    combined.to_excel(fileoutname, header=False, index=False)
    
    for name in excel_names:
        CMD='rm "'+name+'"'
        os.system(CMD)
        
else:
    fileoutname=output_name
    

output_name2=root_dir+"/"+fname+"_RBE_datas.xlsx"
today=str(datetime.now()).replace(' ','_').replace('.','8').replace('.','_')
dtdemande=str(datetime.now().date())
output_name_ax=root_dir+"/"+ident+"_RBE_datas_AX_"+today+".xlsx"

# suppression des excel_names.

    
# traitement du fichier compte-rendu
    
radie='non'
liste_sirens=[]
isradie=""
liste_fj=[]
    
wb = xlrd.open_workbook(fileoutname)
sh=wb.sheet_by_index(0)
for rownum in range(1,sh.nrows):
    
    siren=str(int(sh.row_values(rownum)[0])).zfill(9)
    codefj=str(sh.row_values(rownum)[4])
    isdepotdibe=str(sh.row_values(rownum)[8])
    #print(isdepotdibe)
    if isdepotdibe=='OUI':
        if radie=='oui':
            isradie=str(sh.row_values(rownum)[6])
        if isradie!='OUI':
            liste_sirens.append(siren)
            liste_fj.append(codefj)

book2 = Workbook()
book_ax=Workbook()

feuil1 = book2.active
feuil_ax = book_ax.active

lignes_all=[]
statut_code='nocode'

    
CMD="select prix_unit from ta_clients_ass where id_cptclient='"+ident+"'"
cursor_pg.execute(CMD)
result=cursor_pg.fetchall()
if result:
    prix_unit=result[0][0]
else:
    prix_unit='6.00'


for i in range(len(columns_base_specifique)):
    feuil1.cell(column=i+1,row=1,value=columns_base_specifique[i])
    
nb=len(liste_sirens)/10000
for i in range(int(nb)+1):
    liste_10000=liste_sirens[10000*i:(i+1)*10000]
    liste_fj_10000=liste_fj[10000*i:(i+1)*10000]
    if len(liste_10000) !=0:
        lignes=base_lines(liste_10000,liste_fj_10000,statut_code)
        lignes_all.extend(lignes)
    else:
        break
    
    for j in range(len(lignes_all)):
        for k in range(len(lignes_all[j])):
            feuil1.cell(column=k+1,row=j+2,value=lignes_all[j][k])            
            
    for i in range(len(columns_base_ax)):
        feuil_ax.cell(column=i+1,row=1,value=columns_base_ax[i])
    
    siren_old=""
    index=0
    

    for j in range(len(lignes_all)):
        lignes_ax=[]
        code_article="0069"
        lignes_ax.append(code_article)
        qte=1
        lignes_ax.append(qte)
        unite='Pcs'
        lignes_ax.append(unite)
        prixu=prix_unit
        lignes_ax.append(prixu)
        montant=prixu*qte
        lignes_ax.append(montant)
        greffe=lignes_all[j][2]
        lignes_ax.append(greffe)
        service='DATA'
        lignes_ax.append(service)
        gie=1
        lignes_ax.append(gie)
        siren=str(lignes_all[j][0])
        lignes_ax.append(siren)
        deno=lignes_all[j][1]
        lignes_ax.append(deno)
        #dtdemande=today
        lignes_ax.append(dtdemande)
        
        if siren != siren_old:

        #ligne1=feuil1.row(j+1)
            for k in range(len(lignes_ax)):
                feuil_ax.cell(column=k+1,row=index+2,value=lignes_ax[k])
            siren_old=siren
            index=index+1
    

    # creation du fichier xslx de sortie

# création matérielle du fichier résultant
    book2.save(output_name2)
    book_ax.save(output_name_ax)

cursor_pg.close()
conn_pg.close()